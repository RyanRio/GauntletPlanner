#!/usr/bin/env python3
import argparse
import json
import os
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path, PurePosixPath
import posixpath

import openpyxl

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_DRAW = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"


def parse_workbook(z):
    wb = ET.fromstring(z.read("xl/workbook.xml"))
    sheets = []
    for sheet in wb.find(f"{{{NS_MAIN}}}sheets"):
        name = sheet.attrib.get("name")
        rid = sheet.attrib.get(f"{{{NS_REL}}}id")
        sheets.append((name, rid))
    rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship")}
    return sheets, relmap


def resolve_sheet_path(sheet_name, sheets, relmap):
    for name, rid in sheets:
        if name == sheet_name:
            target = relmap.get(rid)
            if not target:
                raise ValueError(f"No target for sheet {sheet_name}")
            return f"xl/{target}"
    raise ValueError(f"Sheet not found: {sheet_name}")


def get_drawing_path(z, sheet_path):
    sheet_xml = ET.fromstring(z.read(sheet_path))
    drawing = sheet_xml.find(f"{{{NS_MAIN}}}drawing")
    if drawing is None:
        return None
    rid = drawing.attrib.get(f"{{{NS_REL}}}id")
    rels_path = Path(sheet_path).parent / "_rels" / (Path(sheet_path).name + ".rels")
    rels = ET.fromstring(z.read(str(rels_path)))
    relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship")}
    target = relmap.get(rid)
    if not target:
        return None
    base = PurePosixPath(sheet_path).parent
    resolved = PurePosixPath(str(base / PurePosixPath(target)))
    return posixpath.normpath(str(resolved))


def parse_drawings(z, drawing_path):
    if not drawing_path:
        return {}
    drawing_xml = ET.fromstring(z.read(drawing_path))
    rels_path = Path(drawing_path).parent / "_rels" / (Path(drawing_path).name + ".rels")
    try:
        rels = ET.fromstring(z.read(str(rels_path)))
    except KeyError:
        return []
    relmap = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship")}
    base = PurePosixPath(drawing_path).parent

    anchors = []
    for anchor_tag in ("oneCellAnchor", "twoCellAnchor"):
        for anchor in drawing_xml.findall(f"{{{NS_DRAW}}}{anchor_tag}"):
            frm = anchor.find(f"{{{NS_DRAW}}}from")
            if frm is None:
                continue
            col = int(frm.find(f"{{{NS_DRAW}}}col").text)
            row = int(frm.find(f"{{{NS_DRAW}}}row").text)
            blip = anchor.find(f".//{{{NS_A}}}blip")
            if blip is None:
                continue
            embed = blip.attrib.get(f"{{{NS_REL}}}embed")
            target = relmap.get(embed)
            if not target:
                continue
            resolved = posixpath.normpath(str(PurePosixPath(base / PurePosixPath(target))))
            anchors.append((row, col, resolved))
    return anchors


def extract_images(z, anchors, out_dir):
    out_dir.mkdir(parents=True, exist_ok=True)
    mapping = {}
    for row, col, target in anchors:
        target_path = str(PurePosixPath(target))
        if not target_path.startswith("xl/"):
            target_path = str(PurePosixPath("xl") / PurePosixPath(target_path))
        try:
            data = z.read(target_path)
        except KeyError:
            continue
        suffix = Path(target).suffix or ".png"
        out_name = f"r{row+1}_c{col+1}{suffix}"
        out_path = out_dir / out_name
        with open(out_path, "wb") as f:
            f.write(data)
        mapping[(row + 1, col + 1)] = out_name
    return mapping


def detect_headers(ws):
    row1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    row2 = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    headers = []
    for i in range(len(row1)):
        primary = (row2[i] or "").strip() if isinstance(row2[i], str) else row2[i]
        fallback = (row1[i] or "").strip() if isinstance(row1[i], str) else row1[i]
        headers.append(primary or fallback or f"Column {i+1}")
    return headers


def normalize(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def find_boss_names(headers):
    ignore = {"sync pair", "total", "number of solos", "las in-depth infos", "las in depth infos", "readme"}
    return [h for h in headers if h and normalize(h) not in ignore]


def find_boss_sheet(wb, boss_name):
    target = normalize(boss_name)
    for name in wb.sheetnames:
        if target in normalize(name):
            return name
    return None


def parse_boss_sheet(ws):
    header_row = None
    for r in range(1, 40):
        if normalize(ws.cell(r, 1).value) == "who can solo":
            header_row = r
            break
    if header_row is None:
        return {}

    data = {}
    r = header_row + 1
    while r <= ws.max_row:
        name = ws.cell(r, 2).value
        if not name:
            r += 1
            continue
        sync_name = str(name).strip()
        move_level = ws.cell(r, 3).value
        grid = ws.cell(r, 4).value
        min_invest = ws.cell(r, 5).value
        max_invest = ws.cell(r, 6).value
        difficulty = ws.cell(r, 7).value
        notes = ws.cell(r, 8).value
        min_link = ws.cell(r, 5).hyperlink.target if ws.cell(r, 5).hyperlink else None
        max_link = ws.cell(r, 6).hyperlink.target if ws.cell(r, 6).hyperlink else None

        data[sync_name] = {
            "moveLevel": str(move_level).strip() if move_level is not None else "",
            "grid": str(grid).strip() if grid is not None else "",
            "minInvestment": str(min_invest).strip() if min_invest is not None else "",
            "maxInvestment": str(max_invest).strip() if max_invest is not None else "",
            "minVideo": min_link,
            "maxVideo": max_link,
            "difficulty": str(difficulty).strip() if difficulty is not None else "",
            "notes": str(notes).strip() if notes is not None else ""
        }
        r += 1

    return data


def build_nav_image_map(z, wb, nav_sheet_name, out_dir):
    if nav_sheet_name not in wb.sheetnames:
        return {}
    sheets, relmap = parse_workbook(z)
    sheet_path = resolve_sheet_path(nav_sheet_name, sheets, relmap)
    drawing_path = get_drawing_path(z, sheet_path)
    anchors = parse_drawings(z, drawing_path)
    img_map = extract_images(z, anchors, out_dir)

    ws = wb[nav_sheet_name]
    name_by_row = {}
    for r in range(1, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name:
            name_by_row[r] = str(name).strip()

    image_by_name = {}
    for (row, col), filename in img_map.items():
        if col != 1:
            continue
        name = name_by_row.get(row)
        if name:
            image_by_name[name] = filename
    return image_by_name


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--sheet", default="⚜️Current LG(24)")
    parser.add_argument("--nav-sheet", default="💠Navigation")
    parser.add_argument("--out-json", type=Path, default=Path("clears_from_xlsx.json"))
    parser.add_argument("--out-images", type=Path, default=Path("clears_images"))
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}")
    ws = wb[args.sheet]
    headers = detect_headers(ws)
    boss_names = find_boss_names(headers)

    # Determine sync pair column
    sync_idx = None
    for idx, header in enumerate(headers):
        if normalize(header) == "sync pair":
            sync_idx = idx + 1
            break
    if sync_idx is None:
        sync_idx = 1

    # If sync column has no values but next column does, shift
    data_start = 4
    def col_has_values(col):
        for r in range(data_start, min(ws.max_row, data_start + 30)):
            if ws.cell(r, col).value:
                return True
        return False
    if not col_has_values(sync_idx) and col_has_values(sync_idx + 1):
        sync_idx += 1

    ignore = {"sync pair", "total", "number of solos", "las in-depth infos", "las in depth infos", "readme"}
    boss_cols = [i + 1 for i, h in enumerate(headers) if normalize(h) not in ignore and h]

    # Extract images
    boss_details = {}
    with zipfile.ZipFile(args.xlsx, "r") as z:
        image_by_name = build_nav_image_map(z, wb, args.nav_sheet, args.out_images)
        img_map = {}

    for boss in boss_names:
        sheet_name = find_boss_sheet(wb, boss)
        if not sheet_name:
            continue
        boss_details[boss] = parse_boss_sheet(wb[sheet_name])

    rows = []
    for r in range(data_start, ws.max_row + 1):
        name = ws.cell(r, sync_idx).value
        if not name:
            continue
        if normalize(name).startswith("number of solos") or normalize(name) == "readme":
            continue
        row_entry = {
            "row": r,
            "syncPair": str(name).strip(),
            "image": image_by_name.get(str(name).strip()),
            "bosses": {}
        }
        for c in boss_cols:
            boss = headers[c - 1]
            val = ws.cell(r, c).value
            if val is None or val == "":
                continue
            row_entry["bosses"][str(boss)] = str(val).strip()
        rows.append(row_entry)

    payload = {
        "sheet": args.sheet,
        "headers": headers,
        "syncPairColumn": sync_idx,
        "bossColumns": boss_cols,
        "rows": rows,
        "bossDetails": boss_details
    }

    with open(args.out_json, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    print(f"Wrote {args.out_json} and {args.out_images}")


if __name__ == "__main__":
    main()
